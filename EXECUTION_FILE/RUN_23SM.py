import sys
import os
from openpyxl import load_workbook
import pprint

sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
from HIDDEN_FILES.linkData import *
from RENEWAL_UTIL import excel, date

HIDDENFILE_PATH = 링크[4]

# 위탁업체_list = {
#     "COLUMNS": ["E", "I", "J", "N", "R"],  # 가져올 열(알파벳)
#     # 업체명, 시작, 종료, 위탁료(월), 관리비(월)
#     "FILE_NAME": "위탁업체/20220407 위탁업체 list.xlsm",  # 파일명
#     "SHEET_NAME": "21년 서울캠",  # 시트명
#     "FIRST_ROW": 5,  # 시작 행(숫자)
# }

계정별원장자료 = {
    "COLUMNS": ["C", "I", "Q", "T", "R"],  # 가져올 열(알파벳)
    # 전표일자, 대변, 계정과목명, 관리코드명, 적요
    "FILE_NAME": "vendor/계정별원장자료.xlsx",  # 파일명 (xlsx 파일로 변환 필요!!)
    "SHEET_NAME": "1",  # 시트명
    "FIRST_ROW": 2,  # 시작 행(숫자)
}

위탁업체_납부현황_최종 = {
    "COLUMNS": ["B", "D", "E"],
    "FILE_NAME": "vendor/위탁운영업체 납부현황.xlsx",
    "SHEET_NAME": "서울캠",
    "FIRST_ROW": 2,
}


def is_관리비_or_위탁료(적요):
    if "관리비" in 적요:
        return "관리비"
    elif "위탁운영료" in 적요:
        return "위탁료"
    else:
        return False


def read_계정별원장자료():
    COLUMNS, FILE_NAME, SHEET_NAME, FIRST_ROW = 계정별원장자료.values()
    fileName = HIDDENFILE_PATH + FILE_NAME
    data = excel.read_column_data(fileName, SHEET_NAME, COLUMNS, FIRST_ROW)
    data = excel.remove_none_rows(data)

    # 대여료및사용료 > 관리비 | 위탁운영료만 필터
    # [datetime.datetime(2023, 5, 31, 0, 0), 8160000, '대여료및사용료', '(주)텐바이텐', '대학로캠퍼스 관리비(6월)']
    대여료및사용료 = [v for v in data if v[2] == "대여료및사용료"]
    관리비_위탁운영료 = [v for v in 대여료및사용료 if "관리비" in v[4] or "위탁운영료" in v[4]]

    # 예외 관리코드
    exceptions = ["산학협력", "신한은행 홍익대학교지점"]
    excepted = [v for v in 관리비_위탁운영료 if not any(exc in v[3] for exc in exceptions)]

    # 대충 이런형태로 바뀜
    # ['5.31', 8160000, '(주)텐바이텐', '관리비', [6]]
    filtered = [item[:2] + item[3:] for item in excepted]
    parsed_date = [[date.convert_to_string(item[0])] + item[1:] for item in filtered]
    result = [
        [
            v[0],
            v[1],
            v[2],
            is_관리비_or_위탁료(v[3]),
            date.parse_month_part(v[3], len(parsed_date), idx, v),
        ]
        for idx, v in enumerate(parsed_date)
    ]

    return result


def write_위탁업체_납부현황(parsed_data):
    월_행 = [0, "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV"]
    COLUMNS, FILE_NAME, SHEET_NAME, FIRST_ROW = 위탁업체_납부현황_최종.values()
    fileName = HIDDENFILE_PATH + FILE_NAME
    print("위탁업체 납부현황 파일 로드 중...")
    workbook = load_workbook(fileName)
    sheet = workbook[SHEET_NAME]
    print("위탁업체 납부현황 파일 읽는 중...")
    data = excel.read_column_data(fileName, SHEET_NAME, COLUMNS, FIRST_ROW)

    # parsed_data와 excel_data를 비교하여 일치하는 데이터와 해당 인덱스를 포함한 리스트를 반환
    # list: 일치하는 데이터와 해당 인덱스(열)를 포함한 튜플 리스트
    matched_data = []
    for i in range(len(parsed_data)):
        flag = 0
        for j in range(len(data)):
            if (
                data[j][0] in parsed_data[i][2]
                and data[j][1] == parsed_data[i][3]
                and data[j][2] == parsed_data[i][1]
            ):
                matched_data.append((j + 2, parsed_data[i]))
                flag = 1
        if flag == 0:
            print(
                f"\033[95m"
                f"[{i + 1}/{len(parsed_data)}] 매치 실패 - "
                f"{parsed_data[i][2]} {parsed_data[i][3]}"
                f"\033[0m"
            )
        else:
            print(
                f"[{i + 1}/{len(parsed_data)}] 매치 성공 - "
                f"{parsed_data[i][2]} {parsed_data[i][3]}"
            )
    sorted_marched_data = sorted(matched_data, key=lambda x: x[0])

    for row in sorted_marched_data:
        month = row[1][4]

        if len(month) > 1:
            excel.merge_cells_and_input_data(
                sheet,
                f"{월_행[month[0]]}{row[0]}",
                f"{월_행[month[-1]]}{row[0]}",
                row[1][0],
            )
        elif len(month) == 1:
            excel.input_data_to_cell(sheet, f"{월_행[month[0]]}{row[0]}", row[1][0])

    workbook.save(fileName)
    print("저장 완료")


def run():
    result = read_계정별원장자료()
    write_위탁업체_납부현황(result)

if __name__ == '__main__':
    run()
"""
    read_계정별원장자료() : '계정별원장자료' 파일에서 데이터를 가져와 가공
    write_위탁업체_납부현황() : '위탁업체 납부현황' 파일에 가공한 데이터를 작성
"""
