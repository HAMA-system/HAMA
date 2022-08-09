import openpyxl
from FUNC_LIBRARY import errorController
from HIDDEN_FILES import linkData
from HIDDEN_FILES.linkData import *


def load_xls(filename):
    print('파일을 불러오는 중입니다. 잠시만 기다려주세요.')
    try:
        xlsfile = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        return xlsfile
    except:
        errorController.errorMsg(1)
        return None

def load_xls_w(filename):
    print('파일을 불러오는 중입니다. 잠시만 기다려주세요.')
    try:
        xlsfile = openpyxl.load_workbook(filename, read_only=False, data_only=False)
        return xlsfile
    except:
        errorController.errorMsg(1)
        return None

def get_cell_data(file, sheetname ,cell):
    # print(sheetname)
    sheet = file.get_sheet_by_name(sheetname)
    return sheet[cell].value

def get_max_row(file, sheetname, column):
    i = row
    count = 0
    while True:
        cell = column + str(i)

        if get_cell_data(file, sheetname, cell) is None:
            return count
        else :
            count += 1
        i = i + 1

# Z ~ AA 케이스 고려 안함
def get_singleline_data(file, sheetname, firstcell, lastcell):
    cell = firstcell
    data = []

    if firstcell[1:] != lastcell[1:]:
        if 65 <= ord(firstcell[1]) <= 90:
            if firstcell[2:] != lastcell[2:]:
                errorController.errorMsg(2)
                return data
        else:
            errorController.errorMsg(2)
            return data

    while cell != lastcell:
        data.append(get_cell_data(file,sheetname,cell))
        if 65 <= ord(cell[1]) <= 90:
            cell = cell[0] + chr(ord(cell[1])+1) + cell[2:]
        else:
            cell = chr(ord(cell[:1])+1) + cell[1:]
    return data

def all_data_fetch(file, sheetname, firstcell, lastcell):
    fcell = firstcell
    lcell = lastcell

    data = []
    while True:
        cell_data = get_cell_data(file, sheetname, fcell)
        if cell_data is None or cell_data == '_':
            break
        # if cell_data == -1:
        #     continue
        data.append(get_singleline_data(file, sheetname, fcell, lcell))
        if 65 <= ord(fcell[1]) <= 90:
            fcell = fcell[:2] + str(int(fcell[2:]) + 1)
            lcell = lcell[:2] + str(int(lcell[2:]) + 1)
        else:
            fcell = fcell[:1] + str(int(fcell[1:]) + 1)
            lcell = lcell[:1] + str(int(lcell[1:]) + 1)
        # print(fcell + ' ' + lcell)

    return data


def put_cell_data(file, sheetname, cell, text):
    # w = file.active
    # w.cell(row=1,column=1).value
    # s = file[sheetname]
    # s[cell] = text
    s = file[sheetname]
    # print(s)
    if 65 <= ord(cell[1]) <= 90:
        c = (ord(cell[0]) - 64) * 26 + ord(cell[1]) - 64
        r = int(cell[2:])
    else:
        c = ord(cell[0]) - 64
        r = int(cell[1:])

    s.cell(row=r,column=c).value = text

def put_singleline_data(file, sheetname, firstcell, lastcell, line):
    cell = firstcell
    data = []

    if firstcell[1:] != lastcell[1:]:
        if 65 <= ord(firstcell[1]) <= 90:
            if firstcell[2:] != lastcell[2:]:
                errorController.errorMsg(2)
                return data
        else:
            errorController.errorMsg(2)
            return data

    i = 0
    while cell != lastcell:
        put_cell_data(file,sheetname,cell,line[i])
        if 65 <= ord(cell[1]) <= 90:
            cell = cell[0] + chr(ord(cell[1])+1) + cell[2:]
        else:
            cell = chr(ord(cell[:1])+1) + cell[1:]
        i += 1

def put_singleline_data_for_bank(file, sheetname, firstcell, lastcell, line, date):
    cell = firstcell
    data = []
    match_dir = ['E','F','H','I','J','K','M','N','O','P','Q','R','U','V','W','X']
    pass_dir = ['K','R','S']
    if firstcell[1:] != lastcell[1:]:
        if 65 <= ord(firstcell[1]) <= 90:
            if firstcell[2:] != lastcell[2:]:
                errorController.errorMsg(2)
                return data
        else:
            errorController.errorMsg(2)
            return data

    i = 0
    while cell != lastcell:
        if cell[0] in pass_dir:
            if 65 <= ord(cell[1]) <= 90:
                cell = cell[0] + chr(ord(cell[1]) + 1) + cell[2:]
            else:
                cell = chr(ord(cell[:1]) + 1) + cell[1:]
            continue
        if cell[0]=='I':
            put_cell_data(file,sheetname,cell,date)
        elif cell[0] in match_dir:
            ncell = match_dir[i]+cell[1:]
            if line[i] is None:
                tmp = '_'
            else:
                tmp = line[i]
            put_cell_data(file,sheetname,ncell,tmp)

        i += 1

        if 65 <= ord(cell[1]) <= 90:
            cell = cell[0] + chr(ord(cell[1]) + 1) + cell[2:]
        else:
            cell = chr(ord(cell[:1]) + 1) + cell[1:]

def save_xls(file,link):
    print('파일을 저장하는 중입니다. 잠시만 기다려주세요.')
    file.save(link)
    print('파일 저장 완료')

def delete_completed_row(file, sheetname, firstcolumn, lastcolumn, row):
    i = row

    while True:
        firstcell = firstcolumn + str(i)
        lastcell = lastcolumn + str(i)
        cell = firstcell
        # print(str(i) + ': ' + firstcell + lastcell)

        if get_cell_data(file, sheetname, cell) is None:
            break

        if get_cell_data(file, sheetname, cell) == -1:
            print(i)
            while cell != lastcell:
                put_cell_data(file, sheetname, cell, ' ')
                cell = chr(ord(cell[:1]) + 1) + cell[1:]
            # print(str(i)+' deleted')
        i = i + 1

    # save_xls(file)
    # print("all -1 row is deleted")

def get_all_directory_info():
    file = load_xls(linkData.링크[2])
    s1 = list(all_data_fetch(file, '결의내역', 'E15', 'I15'))
    # s2 = list(set(list(map(tuple,all_data_fetch(file, '결의내역(정기)', 'E15', 'G15')))))
    return s1

if __name__ == '__main__':
    # print(get_all_directory_info())
    file = load_xls_w(링크[4]+'afterdata.xlsx')
    put_cell_data(file,'결의내역','H35','test')
    put_cell_data(file,'결의내역','AH25','test')
    put_singleline_data(file,'결의내역','E15','H15',['t','e','st'])
    save_xls(file,링크[4]+'afterdata.xlsx')
    # file = load_xls(linkData.링크[2])
    # target_data = all_data_fetch(file, '결의내역', 'E15', 'G15')
    # print(target_data)
    # for i in range(6):
    #     if target_data[i][1] is not None and target_data[i][1] != '':
    #         isMonthly = True
    #         target_data[i][18] = str(target_data[i][18])
    #         target_data[i][2] = str(target_data[i][2])
    #         # 단일 월
    #         if len(target_data[i][2]) <= 2:
    #             for j in range(len(target_data[i][18])):
    #                 if target_data[i][18][j] == '월':
    #                     l = j - 1
    #                     if j > 1 and 49 <= ord(target_data[i][18][j - 2]) < 58:
    #                         l -= 1
    #                     target_data[i][18] = target_data[i][18].replace(target_data[i][18][l:j], target_data[i][2])
    #             if target_data[i][3] is not None and target_data[i][3] != '':
    #                 target_data[i][3] = str(target_data[i][3])
    #                 for j in range(len(target_data[i][3])):
    #                     if target_data[i][3][j] == '월':
    #                         l = j - 1
    #                         if j > 1 and 49 <= ord(target_data[i][3][j - 2]) < 58:
    #                             l -= 1
    #                         target_data[i][3] = target_data[i][3].replace(target_data[i][3][l:j], target_data[i][2])
    #         # 1,2 월
    #         else:
    #             row = -1
    #             for j in range(len(target_data[i][18])):
    #                 if target_data[i][18][j] == '월':
    #                     l = -1
    #                     r = -1
    #                     for k in range(j-1, -1, -1):
    #                         if 49 <= ord(target_data[i][18][k]) < 58:
    #                             r = k
    #                             break
    #                     if r == -1:
    #                         print(row, "행 적요사항에 월을 찾을 수 없습니다.", sep='')
    #                         break
    #                     for k in range(r-2, -1, -1):
    #                         if 49 <= ord(target_data[i][18][k]) < 58:
    #                             l = k
    #                             if k > 0 and 49 <= ord(target_data[i][18][k - 1]) < 58:
    #                                 l -= 1
    #                             break
    #                     if l == -1:
    #                         print(row, "행 적요사항에 이전 월을 찾을 수 없습니다.", sep='')
    #                         break
    #                     target_data[i][18] = target_data[i][18].replace(target_data[i][18][l:r + 1], target_data[i][2])
    #
    #             if target_data[i][3] is not None and target_data[i][3] != '':
    #                 target_data[i][3] = str(target_data[i][3])
    #                 for j in range(len(target_data[i][3])):
    #                     if target_data[i][3][j] == '월':
    #                         l = -1
    #                         r = -1
    #                         for k in range(j - 1, -1, -1):
    #                             if 49 <= ord(target_data[i][3][k]) < 58:
    #                                 r = k
    #                                 break
    #                         if r == -1:
    #                             print(row, "행 월을 찾을 수 없습니다.", sep='')
    #                             break
    #                         for k in range(r - 2, -1, -1):
    #                             if 49 <= ord(target_data[i][3][k]) < 58:
    #                                 l = k
    #                                 if k > 0 and 49 <= ord(target_data[i][3][k - 1]) < 58:
    #                                     l -= 1
    #                                     break
    #                         if l == -1:
    #                             print(row, "행 이전 월을 찾을 수 없습니다.", sep='')
    #                             break
    #                         target_data[i][3] = target_data[i][3].replace(target_data[i][3][l:r + 1], target_data[i][2])