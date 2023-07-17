import openpyxl
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def read_column_data(filename, sheet_name, columns, first_row):
    """
    Excel 파일에서 지정한 열의 데이터를 읽어옵니다.

    매개변수:
        filename (str): Excel 파일의 경로와 파일명입니다.
        sheet_name (str): Excel 파일에서 읽을 시트의 이름입니다.
        columns (list): 가져올 열의 이름(알파벳)으로 이루어진 리스트입니다.
        first_row (int): 데이터를 읽을 시작 행의 숫자입니다.

    반환값:
        list: 지정한 열의 데이터를 이중배열 형태로 반환합니다.
    """
    workbook = openpyxl.load_workbook(filename, data_only=True)
    sheet = workbook[sheet_name]

    data = []
    for row in range(first_row, sheet.max_row + 1):
        row_data = []
        for column in columns:
            cell_value = sheet[column + str(row)].value
            if type(cell_value) == float:
                cell_value = round(cell_value)
            row_data.append(cell_value)
        data.append(row_data)

    return data


def remove_none_rows(data):
    """
    이중배열에서 하나라도 NoneType인 값이 있는 배열을 제거합니다.

    매개변수:
        data (list): 이중배열 형태의 데이터입니다.

    반환값:
        list: NoneType인 값이 없는 배열로 이루어진 이중배열을 반환합니다.
    """
    return [row for row in data if all(item is not None for item in row)]


def merge_cells_and_input_data(
    sheet: Worksheet, startCellLocation, endCellLocation, data
):
    # 셀 병합
    startCell = sheet[startCellLocation]  # 병합을 시작할 셀
    endCell = sheet[endCellLocation]  # 병합을 종료할 셀

    merged_cell_range = f"{startCellLocation}:{endCellLocation}"
    sheet.merge_cells(merged_cell_range)

    # 병합된 셀에 데이터 입력
    startCell.value = data + "입"
    startCell.alignment = Alignment(horizontal="left", vertical="center")


def input_data_to_cell(sheet: Worksheet, cellLocation, data):
    cell = sheet[cellLocation]
    cell.value = data + "입"
    cell.alignment = Alignment(horizontal="left", vertical="center")
